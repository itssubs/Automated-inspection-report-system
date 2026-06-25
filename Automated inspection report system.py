import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
import os
import sys

#The idea with this project is to create a system capable of inspection, filtering and formatting from a given csv file.

print("="*60)
print("AUTOMATED INSPECTION REPORT SYSTEM")
print("="*60)
print("Required CSV columns:")
print("  Part ID, Dimension, Nominal Value, Actual Value,")
print("  Tolerance, Inspector, Date")
print("="*60)

while True:
    csvlocation = input("Enter the location of the file in the format (C:/....): ")
    csvname = input("Enter the name of the CSV file without extention: ")
    #location verification
    if os.path.exists(f"{csvlocation}/{csvname}.csv"):
        choice_savelocation = input("Save the generated report in the same location (Y/N) ? ")
        while True:
            if choice_savelocation.lower() == 'y':
                savelocation = csvlocation
                break
            elif choice_savelocation.lower() == 'n':
                savelocation = input("Enter the location to save the report and the generated image in the same format as CSV : ")
                if os.path.exists(savelocation):
                    break
                else:
                    print("The path doesn't exist try again.")
            else: 
                print("Enter only y or n.")
        
        #Reading the csv and dataframe initialization
        df = pd.read_csv(f"{csvlocation}/{csvname}.csv")
        requied_cols = ['Part ID', 'Dimension Measured', 'Nominal Value', 'Actual Value', 'Tolerance', 'Inspector']
        missing = [c for c in requied_cols if c not in df.columns]
        if missing:
            print(f"Missing columns: {missing}")
            sys.exit()
        df1 = pd.DataFrame()

        #Creation of summary stats
        df1['Part ID'] = df['Part ID']
        df1['Tolerance'] = df['Tolerance']
        df1['deviation'] = abs(df['Nominal Value'] - df['Actual Value'])
        marginal_value = 0.002
        df1['status'] = np.where((df1['deviation'] > (df['Tolerance'])), 'Fail', np.where(df1['deviation'] > df['Tolerance'] - marginal_value, 'Marginal', 'Pass'))
        mean_dev = df1['deviation'].mean()
        standard_dev = df1['deviation'].std()
        upper_limit = mean_dev + 3* standard_dev
        df1['flag_status'] = np.where(df1['deviation'] > upper_limit, 'flagged', 'Good')

        #merged the datafiles
        merge = pd.merge(df, df1, how = 'outer', on = 'Part ID')

        #Dataframe for the failed parts
        failed_parts = merge[merge['status'] == 'Fail']

        #inspector_summary creation
        inspector_summary = pd.DataFrame()

        inspector_summary = merge.groupby('Inspector').agg(Total_Parts = ('Part ID', 'count'))
        pass_count = merge.groupby('Inspector')['status'].apply(lambda x: (x == 'Pass').sum())
        inspector_summary['Total Passed'] = pass_count
        inspector_summary['Fail Count'] = merge.groupby('Inspector')['status'].apply(lambda x : (x == 'Fail').sum())
        inspector_summary['Marginal Count'] = merge.groupby('Inspector')['status'].apply(lambda x : (x == 'Marginal').sum())
        flagged_count = merge.groupby('Inspector')['flag_status'].apply(lambda x : (x == 'flagged').sum())
        inspector_summary['Flagged count'] = flagged_count

        #Data added to the excel file
        with pd.ExcelWriter(f"{savelocation}/Automated Inspection Report.xlsx") as writer:
            df.to_excel(writer, sheet_name  = 'Raw data', index = False)
            df1.to_excel(writer, sheet_name = 'Summary stats', index = False)
            failed_parts.to_excel(writer, sheet_name = 'Failed Parts Only', index = False)
            inspector_summary.to_excel(writer, sheet_name = 'Inspector Summary')

        #Load the excel file to format the data
        wb = load_workbook(f"{savelocation}/Automated Inspection Report.xlsx")

        #pane freezing and auto adjust column width
        for ws in wb.worksheets:
            ws.freeze_panes = 'A2'
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        max_length = max(max_length , len(str(cell.value)))
                ws.column_dimensions[column_letter].width = max_length + 2

        #Header formatting
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold = True)
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
                cell.fill = PatternFill(fill_type = 'solid', start_color = "00FFFB")

        #color selection for conditional formatting of the status
        green = PatternFill(fill_type = 'solid', start_color = "04FF00", end_color = '04FF00')
        yellow = PatternFill(fill_type = 'solid', start_color = "F6FA00", end_color = 'F6FA00')
        red = PatternFill(fill_type = 'solid', start_color = "FC0000", end_color = 'FC0000')
        ws = wb['Summary stats']

        #Finding the status column and formatting it
        stat_column = None
        for cell in ws[1]:
            if cell.value == 'status':
                stat_column = cell.column
                break

        if stat_column:
            for row_num in range(2, ws.max_row + 1):
                current_cell = ws.cell(row = row_num, column = stat_column)
                value = current_cell.value
                if value == 'Fail':
                    current_cell.fill = red
                elif value == 'Pass':
                    current_cell.fill = green
                else:
                    current_cell.fill = yellow

        else:
            print("The column status couldn't be found")

        print("Data added successfully. Processing the images...")

        #Creation of plots and images
        plt.figure(figsize = (4,4))
        plt.hist(merge['deviation'], bins = 10)
        plt.title('Deviation Distribution')
        plt.xlabel('Deviation')
        plt.ylabel('Frequency')
        plt.tight_layout()
        plt.savefig(f"{savelocation}/deviation_histogram.png", dpi = 300)
        plt.close()

        img = Image(f"{savelocation}/deviation_histogram.png")
        ws.add_image(img, f'A{ws.max_row + 5}')

        pass_fail_counts = merge['status'].value_counts()

        plt.figure(figsize = (3,3))
        plt.pie(pass_fail_counts, labels = pass_fail_counts.index, autopct = '%1.1f%%')
        plt.title('Pass Fail Distribution')
        plt.savefig(f"{savelocation}/pass_fail_pie.png", dpi = 300)
        plt.close()

        img = Image(f"{savelocation}/pass_fail_pie.png")
        ws.add_image(img, f'{get_column_letter(ws.max_column + 5)}2')
        wb.save(f"{savelocation}/Automated Inspection Report.xlsx")
        print('Figure added to the excel sheet successfully.')
        break
    else:
        print("The path cannot be found try again.")
        

        